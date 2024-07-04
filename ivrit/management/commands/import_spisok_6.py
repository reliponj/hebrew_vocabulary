import openpyxl
from django.core.management import BaseCommand

from ivrit.models import Spisok6


def main():
    wb = openpyxl.load_workbook('spisok6.xlsx')
    ws = wb.active

    Spisok6.objects.filter().delete()
    for i in range(2, 1450):
        roots = ws.cell(row=i, column=1).value
        if not roots:
            continue

        words = ws.cell(row=i, column=2).value
        table = ws.cell(row=i, column=3).value
        table_2 = ws.cell(row=i, column=4).value

        vocabulary = Spisok6(
            roots=roots,
            words=words,
            tables=table,
            tables_2=table_2)
        vocabulary.save()


class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        main()
