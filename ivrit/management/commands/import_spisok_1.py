import json

import openpyxl
from django.core.management import BaseCommand

from ivrit.models import Spisok1


def main():
    wb = openpyxl.load_workbook('spisok1.xlsx')
    ws = wb.active

    Spisok1.objects.filter().delete()
    for i in range(2, 67000):
        roots = ws.cell(row=i, column=1).value
        if not roots:
            continue

        words = ws.cell(row=i, column=2).value
        word = ws.cell(row=i, column=3).value
        r = ws.cell(row=i, column=4).value
        links = ws.cell(row=i, column=5).value

        vocabulary = Spisok1(
            roots=roots,
            words=words,
            word=word,
            r=r,
            links=links)
        vocabulary.save()


def excel_to_json():
    wb = openpyxl.load_workbook('spisok1.xlsx')
    ws = wb.active

    data_list = []
    for i in range(2, 67000):
        roots = ws.cell(row=i, column=1).value
        if not roots:
            continue

        words = ws.cell(row=i, column=2).value
        word = ws.cell(row=i, column=3).value
        r = ws.cell(row=i, column=4).value
        links = ws.cell(row=i, column=5).value

        data = {
            "roots": roots,
            "words": words,
            "word": word,
            "r": r,
            "links": links,
        }
        data_list.append(data)

    json_object = json.dumps(data_list, indent=4)

    with open("spisok_1.json", "w", encoding='utf-8') as outfile:
        outfile.write(json_object)


class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        main()
