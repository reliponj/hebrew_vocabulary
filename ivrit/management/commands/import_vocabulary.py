import openpyxl
from django.core.management import BaseCommand

from ivrit.models import Vocabulary


def main():
    wb = openpyxl.load_workbook('slovar.xlsx')
    ws = wb.active

    Vocabulary.objects.filter().delete()
    for i in range(2, 16000):
        root = ws.cell(row=i, column=1).value
        if not root:
            continue

        link = ws.cell(row=i, column=2).value
        binyan = ws.cell(row=i, column=3).value
        word = ws.cell(row=i, column=4).value
        word_u = ws.cell(row=i, column=5).value
        word_a = ws.cell(row=i, column=6).value
        words1 = ws.cell(row=i, column=7).value
        words = ws.cell(row=i, column=8).value

        vocabulary = Vocabulary(
            root=root,
            link=link,
            binyan=binyan,
            word=word,
            word_u=word_u,
            word_a=word_a,
            words1=words1,
            words=words)
        vocabulary.save()


class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        main()
